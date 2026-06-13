import openpyxl
import os
import json
import csv

def sync():
    excel_path = "public/data/productos_inventario.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} no existe.")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    products = []
    
    # Columna 1 (A): ID
    # Columna 2 (B): Nombre
    # Columna 3 (C): Precio (ej. "$3.49" o 3.49)
    # Columna 4 (D): Categoría
    # Columna 5 (E): Subcategoría
    # Columna 6 (F): Imagen (Preview) - ignorar
    # Columna 7 (G): Imagen (Archivo) (ej. "tomates_perita.png" o "tomates_perita")
    # Columna 8 (H): Unidad (ej. "1 Kg")
    # Columna 9 (I): Etiquetas (ej. "Nuevo", "Oferta")
    # Columna 10 (J): Stock
    # Columna 11 (K): Stock Almacén

    for r in range(2, ws.max_row + 1):
        pid = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        
        # Limpiar el precio
        price_val = ws.cell(row=r, column=3).value
        if price_val is None:
            price = 0.0
        elif isinstance(price_val, str):
            try:
                price = float(price_val.replace('$', '').strip())
            except ValueError:
                price = 0.0
        else:
            try:
                price = float(price_val)
            except ValueError:
                price = 0.0

        category = ws.cell(row=r, column=4).value or ""
        subcategory = ws.cell(row=r, column=5).value or ""
        
        img_val = ws.cell(row=r, column=7).value
        if img_val:
            img_val = str(img_val).strip()
            # Si no termina en .png, agregarlo
            if not img_val.endswith('.png'):
                img_val += '.png'
            # Si es solo el nombre de archivo, construir la ruta relativa de la página
            if not img_val.startswith('/'):
                image_path = f"/Ananas/images/products/{img_val}"
            else:
                image_path = img_val
        else:
            # Fallback en caso de que no tenga nombre de archivo de imagen
            fallback_key = str(name).lower().replace(' ', '_').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
            image_path = f"/Ananas/images/products/{fallback_key}.png"

        unit = ws.cell(row=r, column=8).value or ""
        
        labels_val = ws.cell(row=r, column=9).value
        if labels_val:
            labels = [l.strip() for l in str(labels_val).split(',') if l.strip()]
        else:
            labels = []

        stock_val = ws.cell(row=r, column=10).value
        try:
            stock = int(stock_val) if stock_val is not None else 0
        except ValueError:
            stock = 0

        warehouse_val = ws.cell(row=r, column=11).value
        try:
            warehouse_stock = int(warehouse_val) if warehouse_val is not None else 0
        except ValueError:
            warehouse_stock = 0

        products.append({
            'id': str(pid or f"p{r-1}"),
            'name': str(name),
            'price': price,
            'category': str(category),
            'subcategory': str(subcategory),
            'image': image_path,
            'unit': str(unit),
            'labels': labels,
            'stock': stock,
            'warehouseStock': warehouse_stock
        })

    # Recopilar subcategorías únicas por categoría para alimentar el menú dinámico
    category_subs = {}
    for p in products:
        cat_id = p['category']
        sub = p['subcategory']
        if cat_id not in category_subs:
            category_subs[cat_id] = []
        if sub and sub not in category_subs[cat_id]:
            category_subs[cat_id].append(sub)

    categories_base = [
        { 'id': 'frutas-vegetales', 'name': 'Frutas y Vegetales', 'icon': '🍎', 'color': 'bg-red-100 text-red-600' },
        { 'id': 'refrigerados-congelados', 'name': 'Refrigerados', 'icon': '❄️', 'color': 'bg-blue-100 text-blue-600' },
        { 'id': 'viveres', 'name': 'Víveres', 'icon': '🥫', 'color': 'bg-orange-100 text-orange-600' },
        { 'id': 'cuidado-personal-salud', 'name': 'Cuidado Personal', 'icon': '🧴', 'color': 'bg-teal-100 text-teal-600' },
        { 'id': 'limpieza', 'name': 'Limpieza', 'icon': '🧽', 'color': 'bg-cyan-100 text-cyan-600' },
        { 'id': 'licores', 'name': 'Licores', 'icon': '🍷', 'color': 'bg-purple-100 text-purple-600' }
    ]

    for cat in categories_base:
        cat['subcategories'] = category_subs.get(cat['id'], [])

    categories_json = json.dumps(categories_base, indent=2, ensure_ascii=False)
    products_json = json.dumps(products, indent=2, ensure_ascii=False)
    
    ts_content = f"""export interface Product {{
  id: string;
  name: string;
  price: number;
  category: string;
  subcategory: string;
  image: string;
  labels?: string[]; // e.g. "Oferta", "Nuevo"
  unit?: string;     // e.g. "1 Kg", "500g"
  stock?: number;
  warehouseStock?: number;
}}

export interface Category {{
  id: string;
  name: string;
  icon: string;
  color: string;
  subcategories?: string[];
}}

export const categories: Category[] = {categories_json};

export const products: Product[] = {products_json};

export const zones = ['San Luis El Cafetal'];
"""

    with open('src/data/mockDb.ts', 'w', encoding='utf-8') as f:
        f.write(ts_content)
    print("mockDb.ts actualizado exitosamente desde Excel con subcategorías dinámicas.")

    # Generar archivo productos_plantilla.csv
    csv_path = 'public/data/productos_plantilla.csv'
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['id', 'name', 'price', 'category', 'subcategory', 'image', 'unit', 'labels', 'stock', 'warehouseStock'])
        for p in products:
            labels_str = ",".join(p['labels'])
            writer.writerow([
                p['id'], p['name'], f"{p['price']:.2f}", p['category'], 
                p['subcategory'], p['image'], p['unit'], labels_str, 
                p['stock'], p['warehouseStock']
            ])
    print("productos_plantilla.csv actualizado exitosamente desde Excel.")

if __name__ == '__main__':
    sync()
