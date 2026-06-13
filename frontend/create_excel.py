import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment
import random
from PIL import Image as PILImage

# We'll recreate the basic category structure to loop over it
def get_img_path(key):
    path = f"public/images/products/{key}.png"
    if os.path.exists(path):
        return path
    # Fallback to current directory if not found
    path = f"images/products/{key}.png"
    if os.path.exists(path):
        return path
    return None

categories = [
  {
    'id': 'frutas-vegetales',
    'subcategories': ['Enteras', 'Picadas', 'Jugos', 'Frescos', 'Empacados'],
    'products': [
      {'name': 'Tomates Perita', 'sub': 'Frescos', 'unit': '1 Kg', 'price': 3.49, 'img': 'tomates_perita'},
      {'name': 'Cebolla Blanca', 'sub': 'Frescos', 'unit': '1 Kg', 'price': 2.10, 'img': 'cebolla_blanca'},
      {'name': 'Papa Lavada', 'sub': 'Frescos', 'unit': '1 Kg', 'price': 1.80, 'img': 'papa_lavada'},
      {'name': 'Zanahoria', 'sub': 'Frescos', 'unit': '1 Kg', 'price': 1.50, 'img': 'zanahoria'},
      {'name': 'Pimenton Verde', 'sub': 'Frescos', 'unit': '1 Kg', 'price': 4.20, 'img': 'pimenton_verde'},
      {'name': 'Lechuga Romana', 'sub': 'Frescos', 'unit': '1 Unidad', 'price': 1.20, 'img': 'lechuga_romana'},
      {'name': 'Platano Maduro', 'sub': 'Enteras', 'unit': '1 Kg', 'price': 1.90, 'img': 'platano_maduro'},
      {'name': 'Cambur', 'sub': 'Enteras', 'unit': '1 Kg', 'price': 1.10, 'img': 'cambur'},
      {'name': 'Manzana Gala', 'sub': 'Enteras', 'unit': '1 Kg', 'price': 5.50, 'img': 'manzana_gala'},
      {'name': 'Naranja para Jugo', 'sub': 'Enteras', 'unit': '1 Kg', 'price': 1.60, 'img': 'naranja'},
      {'name': 'Lechosa', 'sub': 'Enteras', 'unit': '1 Kg', 'price': 1.75, 'img': 'lechosa'},
      {'name': 'Limones', 'sub': 'Frescos', 'unit': '1 Kg', 'price': 4.19, 'img': 'limones'},
      {'name': 'Pina', 'sub': 'Enteras', 'unit': '1 Unidad', 'price': 2.55, 'img': 'pina'},
      {'name': 'Aguacate', 'sub': 'Frescos', 'unit': '1 Kg', 'price': 5.20, 'img': 'aguacate'},
      {'name': 'Frutas Picadas Mixtas', 'sub': 'Picadas', 'unit': '500g', 'price': 3.00, 'img': 'frutas_picadas'},
    ]
  },
  {
    'id': 'refrigerados-congelados',
    'subcategories': ['Carnes', 'Pollos', 'Embutidos'],
    'products': [
      {'name': 'Carne Molida SV', 'sub': 'Carnes', 'unit': '1 Kg', 'price': 9.95, 'img': 'carne_molida'},
      {'name': 'Bistec de Ganso', 'sub': 'Carnes', 'unit': '1 Kg', 'price': 11.50, 'img': 'bistec_ganso'},
      {'name': 'Costilla de Res', 'sub': 'Carnes', 'unit': '1 Kg', 'price': 7.80, 'img': 'costilla_res'},
      {'name': 'Pechuga de Pollo', 'sub': 'Pollos', 'unit': '1 Kg', 'price': 8.45, 'img': 'pechuga_pollo'},
      {'name': 'Pollo Entero', 'sub': 'Pollos', 'unit': '1 Kg', 'price': 4.50, 'img': 'pollo_entero'},
      {'name': 'Muslos de Pollo', 'sub': 'Pollos', 'unit': '1 Kg', 'price': 5.20, 'img': 'muslos_pollo'},
      {'name': 'Alitas de Pollo', 'sub': 'Pollos', 'unit': '1 Kg', 'price': 6.00, 'img': 'alitas_pollo'},
      {'name': 'Jamon Cocido Plumrose', 'sub': 'Embutidos', 'unit': '500g', 'price': 6.50, 'img': 'jamon_cocido'},
      {'name': 'Queso Amarillo Paisa', 'sub': 'Embutidos', 'unit': '500g', 'price': 7.20, 'img': 'queso_amarillo'},
      {'name': 'Queso Blanco Duro', 'sub': 'Embutidos', 'unit': '1 Kg', 'price': 6.80, 'img': 'queso_blanco'},
      {'name': 'Salchichas Plumrose', 'sub': 'Embutidos', 'unit': '1 Paquete', 'price': 4.50, 'img': 'salchichas'},
      {'name': 'Tocino Ahumado', 'sub': 'Embutidos', 'unit': '250g', 'price': 4.80, 'img': 'tocino'},
      {'name': 'Chorizo Carupanero', 'sub': 'Embutidos', 'unit': '500g', 'price': 5.50, 'img': 'chorizo'},
      {'name': 'Queso Guayanes', 'sub': 'Embutidos', 'unit': '500g', 'price': 4.50, 'img': 'queso_guayanes'},
      {'name': 'Nuggets Congelados', 'sub': 'Pollos', 'unit': '1 Paquete', 'price': 5.99, 'img': 'nuggets'},
    ]
  },
  {
    'id': 'viveres',
    'subcategories': ['Arroz', 'Pasta', 'Enlatados', 'Refrescos'],
    'products': [
      {'name': 'Coca Cola Original', 'sub': 'Refrescos', 'unit': '2 L', 'price': 2.50, 'img': 'coca_cola'},
      {'name': 'Pepsi Cola', 'sub': 'Refrescos', 'unit': '2 L', 'price': 2.20, 'img': 'pepsi'},
      {'name': 'Chinotto', 'sub': 'Refrescos', 'unit': '2 L', 'price': 2.20, 'img': 'chinotto'},
      {'name': 'Frescolita', 'sub': 'Refrescos', 'unit': '2 L', 'price': 2.20, 'img': 'frescolita'},
      {'name': 'Maltin Polar', 'sub': 'Refrescos', 'unit': '1.5 L', 'price': 2.00, 'img': 'maltin'},
      {'name': 'Arroz Blanco Mary', 'sub': 'Arroz', 'unit': '1 Kg', 'price': 1.20, 'img': 'arroz_blanco'},
      {'name': 'Arroz Primor', 'sub': 'Arroz', 'unit': '1 Kg', 'price': 1.30, 'img': 'arroz_primor'},
      {'name': 'Pasta Capri Espagueti', 'sub': 'Pasta', 'unit': '1 Kg', 'price': 1.60, 'img': 'pasta_capri'},
      {'name': 'Harina PAN', 'sub': 'Arroz', 'unit': '1 Kg', 'price': 1.15, 'img': 'harina_pan'},
      {'name': 'Caraotas Negras Mary', 'sub': 'Enlatados', 'unit': '500g', 'price': 1.80, 'img': 'caraotas'},
      {'name': 'Lentejas', 'sub': 'Enlatados', 'unit': '500g', 'price': 2.10, 'img': 'lentejas'},
      {'name': 'Atun Margarita en Aceite', 'sub': 'Enlatados', 'unit': '140g', 'price': 2.50, 'img': 'atun_aceite'},
      {'name': 'Atun Margarita en Agua', 'sub': 'Enlatados', 'unit': '140g', 'price': 2.50, 'img': 'atun_agua'},
      {'name': 'Maiz Dulce en Lata', 'sub': 'Enlatados', 'unit': '300g', 'price': 1.50, 'img': 'maiz_lata'},
      {'name': 'Aceite de Maiz Mazeite', 'sub': 'Arroz', 'unit': '1 Litro', 'price': 3.50, 'img': 'aceite_maiz'},
    ]
  },
  {
    'id': 'cuidado-personal-salud',
    'subcategories': ['Aseo', 'Farmacia'],
    'products': [
      {'name': 'Jabon Protex', 'sub': 'Aseo', 'unit': '3 Unidades', 'price': 3.50, 'img': 'jabon_protex'},
      {'name': 'Desodorante Dove', 'sub': 'Aseo', 'unit': '50ml', 'price': 4.20, 'img': 'desodorante'},
      {'name': 'Crema Dental Colgate', 'sub': 'Aseo', 'unit': '100g', 'price': 2.80, 'img': 'crema_dental'},
      {'name': 'Shampoo Pantene', 'sub': 'Aseo', 'unit': '400ml', 'price': 6.50, 'img': 'shampoo'},
      {'name': 'Papel Higienico Scott', 'sub': 'Aseo', 'unit': '4 Rollos', 'price': 3.80, 'img': 'papel_higienico'},
      {'name': 'Toallas Sanitarias Always', 'sub': 'Aseo', 'unit': '10 Unidades', 'price': 2.90, 'img': 'toallas_sanitarias'},
      {'name': 'Maquina de Afeitar Gillette', 'sub': 'Aseo', 'unit': '2 Unidades', 'price': 3.00, 'img': 'maquina_afeitar'},
      {'name': 'Crema Corporal Lubriderm', 'sub': 'Aseo', 'unit': '400ml', 'price': 8.50, 'img': 'crema_corporal'},
      {'name': 'Algodon en Motas', 'sub': 'Farmacia', 'unit': '100g', 'price': 1.50, 'img': 'algodon'},
      {'name': 'Hisopos Q-Tips', 'sub': 'Farmacia', 'unit': '100 Unidades', 'price': 1.80, 'img': 'hisopos'},
      {'name': 'Alcohol Isopropilico', 'sub': 'Farmacia', 'unit': '250ml', 'price': 2.20, 'img': 'alcohol'},
      {'name': 'Agua Oxigenada', 'sub': 'Farmacia', 'unit': '100ml', 'price': 1.00, 'img': 'agua_oxigenada'},
      {'name': 'Curitas Band-Aid', 'sub': 'Farmacia', 'unit': 'Caja', 'price': 3.00, 'img': 'curitas'},
    ]
  },
  {
    'id': 'limpieza',
    'subcategories': ['Detergentes', 'Accesorios'],
    'products': [
      {'name': 'Detergente Liquido Ariel', 'sub': 'Detergentes', 'unit': '2 L', 'price': 8.50, 'img': 'detergente_liquido'},
      {'name': 'Detergente en Polvo ACE', 'sub': 'Detergentes', 'unit': '1 Kg', 'price': 4.50, 'img': 'detergente_polvo'},
      {'name': 'Suavizante Suavitel', 'sub': 'Detergentes', 'unit': '1 L', 'price': 3.80, 'img': 'suavizante'},
      {'name': 'Lavaplatos Liquido Las Llaves', 'sub': 'Detergentes', 'unit': '500ml', 'price': 2.50, 'img': 'lavaplatos_liquido'},
      {'name': 'Lavaplatos en Crema Axion', 'sub': 'Detergentes', 'unit': '250g', 'price': 1.80, 'img': 'lavaplatos_crema'},
      {'name': 'Cloro Nevex', 'sub': 'Detergentes', 'unit': '1 L', 'price': 1.50, 'img': 'cloro'},
      {'name': 'Desinfectante Mistolin', 'sub': 'Detergentes', 'unit': '1 L', 'price': 2.20, 'img': 'desinfectante'},
      {'name': 'Esponja Scotch-Brite', 'sub': 'Accesorios', 'unit': '3 Unidades', 'price': 3.50, 'img': 'esponja'},
      {'name': 'Mopa de Algodon', 'sub': 'Accesorios', 'unit': '1 Unidad', 'price': 5.00, 'img': 'mopa'},
      {'name': 'Escoba con Mango', 'sub': 'Accesorios', 'unit': '1 Unidad', 'price': 4.50, 'img': 'escoba'},
      {'name': 'Coleto Tradicional', 'sub': 'Accesorios', 'unit': '1 Unidad', 'price': 2.00, 'img': 'coleto'},
      {'name': 'Bolsas de Basura Grandes', 'sub': 'Accesorios', 'unit': '10 Unidades', 'price': 3.20, 'img': 'bolsas_basura'},
    ]
  },
  {
    'id': 'licores',
    'subcategories': ['Vinos', 'Cervezas', 'Destilados'],
    'products': [
      {'name': 'Cerveza Polar Pilsen', 'sub': 'Cervezas', 'unit': 'Caja 36', 'price': 25.00, 'img': 'cerveza_pilsen'},
      {'name': 'Cerveza Polar Light', 'sub': 'Cervezas', 'unit': 'Caja 36', 'price': 25.00, 'img': 'cerveza_light'},
      {'name': 'Cerveza Solera Verde', 'sub': 'Cervezas', 'unit': '6 Pack', 'price': 6.50, 'img': 'cerveza_solera'},
      {'name': 'Cerveza Zulia', 'sub': 'Cervezas', 'unit': '6 Pack', 'price': 6.00, 'img': 'cerveza_zulia'},
      {'name': 'Ron Santa Teresa Linaje', 'sub': 'Destilados', 'unit': '750ml', 'price': 18.00, 'img': 'ron_teresa'},
      {'name': 'Ron Cacique Anejo', 'sub': 'Destilados', 'unit': '750ml', 'price': 12.00, 'img': 'ron_cacique'},
      {'name': 'Ron Pampero Aniversario', 'sub': 'Destilados', 'unit': '750ml', 'price': 22.00, 'img': 'ron_pampero'},
      {'name': 'Ron Diplomatico Reserva Exclusiva', 'sub': 'Destilados', 'unit': '750ml', 'price': 35.00, 'img': 'ron_diplomatico'},
      {'name': 'Whisky Buchanan\'s 12 Anos', 'sub': 'Destilados', 'unit': '750ml', 'price': 42.00, 'img': 'whisky_buchanans'},
      {'name': 'Whisky Old Parr 12 Anos', 'sub': 'Destilados', 'unit': '750ml', 'price': 38.00, 'img': 'whisky_parr'},
      {'name': 'Vodka Gordon\'s', 'sub': 'Destilados', 'unit': '750ml', 'price': 10.00, 'img': 'vodka'},
      {'name': 'Vino Tinto Casillero del Diablo', 'sub': 'Vinos', 'unit': '750ml', 'price': 12.50, 'img': 'vino_tinto'},
      {'name': 'Vino Blanco Santa Helena', 'sub': 'Vinos', 'unit': '750ml', 'price': 8.50, 'img': 'vino_blanco'},
    ]
  }
]

def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ['ID', 'Nombre', 'Precio', 'Categoría', 'Subcategoría', 'Imagen (Preview)', 'Imagen (Archivo)', 'Unidad', 'Etiquetas', 'Stock', 'Stock Almacén']
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20 # Image column
    ws.column_dimensions['G'].width = 25 # Image filename column
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15

    id_counter = 1
    current_row = 2

    # Create temporary thumbs dir inside workspace
    temp_thumbs_dir = "public/data/temp_thumbs"
    os.makedirs(temp_thumbs_dir, exist_ok=True)

    for cat in categories:
        for prod in cat['products']:
            stock = random.randint(10, 100)
            warehouse_stock = random.randint(50, 500)
            
            label = ''
            r = random.random()
            if r > 0.9:
                label = 'Nuevo'
            elif r > 0.8:
                label = 'Oferta'
                
            # Set row height to accommodate image
            ws.row_dimensions[current_row].height = 60

            # Insert text data
            ws.cell(row=current_row, column=1, value=f"p{id_counter}")
            ws.cell(row=current_row, column=2, value=prod['name'])
            ws.cell(row=current_row, column=3, value=f"${prod['price']:.2f}")
            ws.cell(row=current_row, column=4, value=cat['id'])
            ws.cell(row=current_row, column=5, value=prod['sub'])
            ws.cell(row=current_row, column=7, value=f"{prod['img']}.png")
            ws.cell(row=current_row, column=8, value=prod['unit'])
            ws.cell(row=current_row, column=9, value=label)
            ws.cell(row=current_row, column=10, value=stock)
            ws.cell(row=current_row, column=11, value=warehouse_stock)

            # Insert Image
            img_path = get_img_path(prod['img'])
            if img_path:
                try:
                    # Open and resize image to fit in cell
                    with PILImage.open(img_path) as pil_img:
                        pil_img.thumbnail((70, 70))
                        temp_path = f"{temp_thumbs_dir}/thumb_{prod['img']}.png"
                        pil_img.save(temp_path)
                        
                        xl_img = ExcelImage(temp_path)
                        
                        # Position image in column F
                        cell_address = f"F{current_row}"
                        ws.add_image(xl_img, cell_address)
                except Exception as e:
                    print(f"Error loading image {img_path}: {e}")
                    ws.cell(row=current_row, column=6, value="(Img no encontrada)")
            else:
                ws.cell(row=current_row, column=6, value="(Sin img)")

            for col in range(1, 12):
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center')

            id_counter += 1
            current_row += 1

    file_path = "public/data/productos_inventario.xlsx"
    wb.save(file_path)
    print(f"Excel guardado en: {file_path}")

    # Clean up local temporary thumbnails folder
    try:
        import shutil
        if os.path.exists(temp_thumbs_dir):
            shutil.rmtree(temp_thumbs_dir)
    except Exception as e:
        print(f"Error cleaning up temp thumbs: {e}")

if __name__ == "__main__":
    create_excel()
