import pandas as pd
import openpyxl
import os
import subprocess

EXCEL_FILE = "Inventario_Ananas.xlsx"

def print_header(title):
    print("\n" + "="*50)
    print(f"🍍 {title}")
    print("="*50)

def main():
    print_header("Procesador de Pedidos y Restock Automático")
    
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Error: No se encuentra el archivo {EXCEL_FILE}")
        return

    # Load Excel file with openpyxl to keep formatting if possible, 
    # but pandas is easier to read. We will use openpyxl directly to update cells safely.
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find column indices
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    
    id_col = headers.get('ID')
    name_col = headers.get('Nombre')
    stock_col = headers.get('Stock')
    warehouse_col = headers.get('Stock Almacén')

    if not all([id_col, name_col, stock_col, warehouse_col]):
        print("❌ Error: Faltan columnas requeridas en el Excel (ID, Nombre, Stock, Stock Almacén).")
        return

    print("Ingrese los productos vendidos. Deje el ID en blanco para terminar y procesar.")
    
    orders = []
    while True:
        prod_id = input("\n👉 ID del producto (ej. p1) o presione Enter para procesar: ").strip()
        if not prod_id:
            break
            
        try:
            qty = int(input(f"   Cantidad vendida de {prod_id}: ").strip())
            orders.append((prod_id, qty))
        except ValueError:
            print("   ❌ Cantidad inválida, intente de nuevo.")

    if not orders:
        print("No se ingresaron productos. Saliendo...")
        return

    print_header("Procesando Inventario")
    
    # Process each order
    updated = False
    for prod_id, qty in orders:
        found = False
        for row in range(2, ws.max_row + 1):
            cell_id = ws.cell(row=row, column=id_col).value
            if str(cell_id).strip() == str(prod_id).strip():
                found = True
                name = ws.cell(row=row, column=name_col).value
                current_stock = int(ws.cell(row=row, column=stock_col).value or 0)
                current_warehouse = int(ws.cell(row=row, column=warehouse_col).value or 0)
                
                print(f"\n📦 Procesando: {name} (ID: {prod_id}) - Vendidos: {qty}")
                print(f"   Stock en tienda actual: {current_stock} | Stock en almacén actual: {current_warehouse}")
                
                # Logic: Deduct from stock first
                new_stock = current_stock - qty
                new_warehouse = current_warehouse
                
                # If stock drops below a safe threshold (e.g. 5) or becomes negative, replenish from warehouse
                needed = 0
                if new_stock < 5:
                    # We want to replenish to at least 15 if possible
                    needed = 15 - new_stock
                    
                if needed > 0 and new_warehouse > 0:
                    transfer = min(needed, new_warehouse)
                    new_stock += transfer
                    new_warehouse -= transfer
                    print(f"   🔄 Reponiendo desde almacén: se pasaron {transfer} unidades a tienda.")
                
                if new_stock < 0:
                    print(f"   ⚠️ ADVERTENCIA: El stock de {name} quedó en negativo ({new_stock}).")

                ws.cell(row=row, column=stock_col).value = new_stock
                ws.cell(row=row, column=warehouse_col).value = new_warehouse
                
                print(f"   ✅ Nuevo Stock Tienda: {new_stock} | Nuevo Stock Almacén: {new_warehouse}")
                updated = True
                break
                
        if not found:
            print(f"\n❌ Error: No se encontró el producto con ID '{prod_id}' en el inventario.")

    if updated:
        wb.save(EXCEL_FILE)
        print_header("Actualización Exitosa")
        print(f"✅ Se ha actualizado '{EXCEL_FILE}'.")
        print("🔄 Sincronizando catálogo para la página web...")
        
        try:
            subprocess.run(["python3", "sync_excel_to_db.py"], check=True)
            print("✅ Base de datos de la web actualizada con éxito.")
        except Exception as e:
            print(f"❌ Error al sincronizar con la web: {e}")
            print("Ejecute 'python3 sync_excel_to_db.py' manualmente.")
    else:
        print("\nNo se realizaron cambios en el inventario.")

if __name__ == "__main__":
    main()
